#!/usr/bin/env python3
"""Tests for stm_to_excel.py.

Unit tests for transform translation (no CLI dependency) and integration
tests against the canonical example corpus.
"""

from __future__ import annotations

import subprocess
from pathlib import Path

import pytest

# Ensure imports work when run from repo root
SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent.parent.parent

# Import the module under test by adding its directory to sys.path first.
# The sys import and stm_to_excel import must follow the path manipulation,
# so E402 (module-level import not at top) is expected here.
import sys  # noqa: E402
sys.path.insert(0, str(SCRIPT_DIR))
import stm_to_excel  # noqa: E402


# ── Transform translation unit tests ────────────────────────────────────

class TestTranslateSingleToken:
    """Test _translate_single_token for individual pipeline steps."""

    def test_trim(self):
        assert stm_to_excel._translate_single_token("trim") == "trim whitespace"

    def test_lowercase(self):
        assert stm_to_excel._translate_single_token("lowercase") == "convert to lowercase"

    def test_uppercase(self):
        assert stm_to_excel._translate_single_token("uppercase") == "convert to uppercase"

    def test_title_case(self):
        assert stm_to_excel._translate_single_token("title_case") == "convert to title case"

    def test_null_if_empty(self):
        assert stm_to_excel._translate_single_token("null_if_empty") == "set null if empty"

    def test_validate_email(self):
        assert stm_to_excel._translate_single_token("validate_email") == "validate email format"

    def test_now_utc(self):
        assert stm_to_excel._translate_single_token("now_utc()") == "current UTC timestamp"

    def test_to_iso8601(self):
        assert stm_to_excel._translate_single_token("to_iso8601") == "format as ISO 8601"

    def test_coalesce(self):
        assert stm_to_excel._translate_single_token("coalesce(0)") == "default to 0 if null"

    def test_round(self):
        assert stm_to_excel._translate_single_token("round(2)") == "round to 2 decimal places"

    def test_round_bare(self):
        assert stm_to_excel._translate_single_token("round") == "round to nearest integer"

    def test_truncate(self):
        assert stm_to_excel._translate_single_token("truncate(5000)") == "truncate to 5000 characters"

    def test_parse(self):
        assert stm_to_excel._translate_single_token('parse("MM/DD/YYYY")') == 'parse as "MM/DD/YYYY"'

    def test_encrypt(self):
        result = stm_to_excel._translate_single_token("encrypt(AES-256-GCM, secrets.tax_encryption_key)")
        assert result == "encrypt (AES-256-GCM)"

    def test_uuid_v5(self):
        result = stm_to_excel._translate_single_token('uuid_v5("6ba7b810-9dad-11d1-80b4-00c04fd430c8", CUST_ID)')
        assert result == "generate UUID v5"

    def test_multiply(self):
        assert stm_to_excel._translate_single_token("* 100") == "multiply by 100"

    def test_nl_string(self):
        result = stm_to_excel._translate_single_token('"Extract digits and format as E.164"')
        assert result == "Extract digits and format as E.164"

    def test_unknown_token_passthrough(self):
        assert stm_to_excel._translate_single_token("custom_func") == "custom_func"

    def test_empty(self):
        assert stm_to_excel._translate_single_token("") == ""

    def test_escape_html(self):
        assert stm_to_excel._translate_single_token("escape_html") == "escape HTML characters"

    def test_drop_if_invalid(self):
        assert stm_to_excel._translate_single_token("drop_if_invalid") == "drop if invalid"

    def test_error_if_null(self):
        assert stm_to_excel._translate_single_token("error_if_null") == "error if null"

    def test_assume_utc(self):
        assert stm_to_excel._translate_single_token("assume_utc") == "assume UTC timezone"


class TestParseMapBlock:
    """Test _parse_map_block for map { } expressions."""

    def test_simple_map(self):
        raw = 'map {\n      R: "retail"\n      B: "business"\n    }'
        inline, entries, is_cond = stm_to_excel._parse_map_block(raw)
        assert not is_cond
        assert len(entries) == 2
        assert 'R = "retail"' in inline
        assert 'B = "business"' in inline

    def test_conditional_map(self):
        raw = 'map {\n      < 1000: "bronze"\n      < 5000: "silver"\n      default: "platinum"\n    }'
        inline, entries, is_cond = stm_to_excel._parse_map_block(raw)
        assert is_cond
        assert len(entries) == 3
        assert "Conditional" in inline

    def test_null_key(self):
        raw = 'map {\n      null: "retail"\n    }'
        inline, entries, is_cond = stm_to_excel._parse_map_block(raw)
        assert not is_cond
        assert '(empty) = "retail"' in inline

    def test_default_key_in_simple(self):
        raw = 'map {\n      _: "unknown"\n    }'
        inline, entries, is_cond = stm_to_excel._parse_map_block(raw)
        assert not is_cond
        assert '(other) = "unknown"' in inline


class TestTranslateTransform:
    """Test translate_transform for full pipeline expressions."""

    def test_pure_nl(self):
        human, entries, cond = stm_to_excel.translate_transform(
            None, '"Extract all digits and format"', "nl",
        )
        assert human == "Extract all digits and format"
        assert not entries
        assert not cond

    def test_structural_chain(self):
        human, entries, cond = stm_to_excel.translate_transform(
            ["trim", "lowercase", "validate_email", "null_if_invalid"],
            None, "structural",
        )
        assert human == "trim whitespace \u2192 convert to lowercase \u2192 validate email format \u2192 set null if invalid"
        assert not cond

    def test_mixed(self):
        human, entries, cond = stm_to_excel.translate_transform(
            ["warn_if_invalid"],
            '"Extract all digits. Format as E.164."',
            "mixed",
        )
        assert "warn if invalid" in human
        assert "Extract all digits" in human
        assert "\u2192" in human

    def test_no_transform(self):
        human, entries, cond = stm_to_excel.translate_transform(
            None, None, "none",
        )
        assert human == ""

    def test_map_simple(self):
        raw = 'map {\n      A: "active"\n      S: "suspended"\n    }'
        human, entries, cond = stm_to_excel.translate_transform(
            [raw], None, "structural",
        )
        assert 'A = "active"' in human
        assert not cond

    def test_map_conditional(self):
        raw = 'map {\n      < 1000: "bronze"\n      default: "platinum"\n    }'
        human, entries, cond = stm_to_excel.translate_transform(
            [raw], None, "structural",
        )
        assert cond
        assert "Conditional" in human
        assert len(entries) == 2


# ── Integration tests (require satsuma CLI) ──────────────────────────────

def _skip_if_no_cli():
    """Skip test if satsuma CLI is not available."""
    try:
        result = subprocess.run(
            ["npx", "satsuma", "--version"],
            capture_output=True, text=True, timeout=10,
        )
        if result.returncode != 0:
            pytest.skip("satsuma CLI not available")
    except FileNotFoundError:
        pytest.skip("npx not available")


def _run_stm_to_excel(stm_files: list[str], output: str, extra_args: list[str] | None = None) -> str:
    """Run stm_to_excel.py and return stdout."""
    cmd = [
        "python3", str(SCRIPT_DIR / "stm_to_excel.py"),
        *stm_files, "-o", output,
        "--timestamp", "2026-01-01T00:00:00Z",
    ]
    if extra_args:
        cmd.extend(extra_args)
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
    if result.returncode != 0:
        pytest.fail(f"stm_to_excel failed: {result.stderr}")
    return result.stdout


class TestIntegrationDbToDb:
    """Integration tests against examples/db-to-db.stm."""

    @pytest.fixture(autouse=True)
    def setup(self, tmp_path):
        _skip_if_no_cli()
        self.output = str(tmp_path / "db-to-db.xlsx")
        self.stm = str(REPO_ROOT / "examples" / "db-to-db.stm")
        _run_stm_to_excel([self.stm], self.output)
        import openpyxl
        self.wb = openpyxl.load_workbook(self.output)

    def test_tab_names(self):
        names = self.wb.sheetnames
        assert "Overview" in names
        assert "Issues" in names
        assert any(n.startswith("Map") for n in names)
        assert any(n.startswith("Tgt") for n in names)
        assert any(n.startswith("Src") for n in names)

    def test_overview_title(self):
        ws = self.wb["Overview"]
        assert ws.cell(1, 1).value == "Legacy Customer Migration"

    def test_issues_count(self):
        ws = self.wb["Issues"]
        # 6 warnings, each on its own row starting at row 2
        count = sum(1 for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value)
        assert count == 6

    def test_issues_types(self):
        ws = self.wb["Issues"]
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 3).value:
                assert ws.cell(r, 3).value in ("Warning", "Question")

    def test_mapping_arrow_count(self):
        ws = None
        for name in self.wb.sheetnames:
            if name.startswith("Map"):
                ws = self.wb[name]
                break
        assert ws is not None
        # Count data rows (after header, excluding note row)
        arrows = sum(
            1 for r in range(1, ws.max_row + 1)
            if ws.cell(r, 4).value == "\u2192" and isinstance(ws.cell(r, 1).value, int)
        )
        assert arrows == 19

    def test_computed_field_dash(self):
        """Computed/derived fields should show '—' in source column."""
        ws = None
        for name in self.wb.sheetnames:
            if name.startswith("Map"):
                ws = self.wb[name]
                break
        # Find a derived field (display_name, address_id, or migration_timestamp)
        found = False
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 2).value == "\u2014":  # em dash
                found = True
                break
        assert found, "No computed field with em-dash found"

    def test_schema_tab_fields(self):
        ws = self.wb["Tgt - postgres_db"]
        # Should have 19 fields (from summary) + header row + note row
        field_count = sum(
            1 for r in range(1, ws.max_row + 1)
            if isinstance(ws.cell(r, 1).value, int)
        )
        assert field_count >= 18  # at least 18 fields

    def test_schema_pk_marked(self):
        ws = self.wb["Tgt - postgres_db"]
        found_pk = False
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 4).value == "Yes":  # PK column
                found_pk = True
                break
        assert found_pk

    def test_header_styling(self):
        """Header rows should have dark charcoal background."""
        ws = self.wb["Issues"]
        fill = ws.cell(1, 1).fill
        assert fill.start_color.rgb is not None

    def test_lookup_tabs_exist(self):
        ref_tabs = [n for n in self.wb.sheetnames if n.startswith("Ref")]
        assert len(ref_tabs) >= 1

    def test_snapshot_warning(self):
        ws = self.wb["Overview"]
        found = False
        for r in range(1, ws.max_row + 1):
            val = ws.cell(r, 1).value or ""
            if "read-only snapshot" in val.lower():
                found = True
                break
        assert found


class TestIntegrationSfdc:
    """Integration tests against examples/sfdc_to_snowflake.stm."""

    @pytest.fixture(autouse=True)
    def setup(self, tmp_path):
        _skip_if_no_cli()
        self.output = str(tmp_path / "sfdc.xlsx")
        self.stm = str(REPO_ROOT / "examples" / "sfdc_to_snowflake.stm")
        _run_stm_to_excel([self.stm], self.output)
        import openpyxl
        self.wb = openpyxl.load_workbook(self.output)

    def test_tab_count(self):
        # Overview + Issues + 1 mapping + 3 schemas + lookups
        assert len(self.wb.sheetnames) >= 5

    def test_multi_source(self):
        """Mapping with multiple sources should still work."""
        ws = None
        for name in self.wb.sheetnames:
            if name.startswith("Map"):
                ws = self.wb[name]
                break
        assert ws is not None
        arrows = sum(
            1 for r in range(1, ws.max_row + 1)
            if ws.cell(r, 4).value == "\u2192" and isinstance(ws.cell(r, 1).value, int)
        )
        assert arrows == 10

    def test_overview_title(self):
        ws = self.wb["Overview"]
        assert ws.cell(1, 1).value == "Salesforce to Snowflake Pipeline"

    def test_issues_has_warning(self):
        ws = self.wb["Issues"]
        assert ws.cell(2, 3).value == "Warning"


class TestIntegrationOptions:
    """Test CLI option behaviour."""

    @pytest.fixture(autouse=True)
    def setup(self):
        _skip_if_no_cli()
        self.stm = str(REPO_ROOT / "examples" / "db-to-db.stm")

    def test_no_issues_flag(self, tmp_path):
        output = str(tmp_path / "no-issues.xlsx")
        _run_stm_to_excel([self.stm], output, ["--no-issues"])
        import openpyxl
        wb = openpyxl.load_workbook(output)
        assert "Issues" not in wb.sheetnames

    def test_no_schemas_flag(self, tmp_path):
        output = str(tmp_path / "no-schemas.xlsx")
        _run_stm_to_excel([self.stm], output, ["--no-schemas"])
        import openpyxl
        wb = openpyxl.load_workbook(output)
        assert not any(n.startswith("Src") or n.startswith("Tgt") for n in wb.sheetnames)

    def test_title_override(self, tmp_path):
        output = str(tmp_path / "titled.xlsx")
        _run_stm_to_excel([self.stm], output, ["--title", "Custom Title"])
        import openpyxl
        wb = openpyxl.load_workbook(output)
        assert wb["Overview"].cell(1, 1).value == "Custom Title"

    def test_reproducibility(self, tmp_path):
        """Two runs with same timestamp produce identical cell content."""
        out1 = str(tmp_path / "run1.xlsx")
        out2 = str(tmp_path / "run2.xlsx")
        _run_stm_to_excel([self.stm], out1)
        _run_stm_to_excel([self.stm], out2)
        import openpyxl
        wb1 = openpyxl.load_workbook(out1)
        wb2 = openpyxl.load_workbook(out2)
        assert wb1.sheetnames == wb2.sheetnames
        for name in wb1.sheetnames:
            ws1 = wb1[name]
            ws2 = wb2[name]
            assert ws1.max_row == ws2.max_row, f"Row count mismatch in {name}"
            assert ws1.max_column == ws2.max_column, f"Col count mismatch in {name}"
            for r in range(1, (ws1.max_row or 0) + 1):
                for c in range(1, (ws1.max_column or 0) + 1):
                    v1 = ws1.cell(r, c).value
                    v2 = ws2.cell(r, c).value
                    assert v1 == v2, f"Cell {name}!{r},{c}: {v1!r} != {v2!r}"


class TestIntegrationFragments:
    """Test handling of files with fragments."""

    @pytest.fixture(autouse=True)
    def setup(self, tmp_path):
        _skip_if_no_cli()
        self.output = str(tmp_path / "sfdc-frags.xlsx")
        self.stm = str(REPO_ROOT / "examples" / "sfdc_to_snowflake.stm")
        _run_stm_to_excel([self.stm], self.output)
        import openpyxl
        self.wb = openpyxl.load_workbook(self.output)

    def test_fragment_notes_in_schema(self):
        """Schema tabs should mention fragment origins in notes."""
        for name in self.wb.sheetnames:
            if name.startswith("Src") or name.startswith("Tgt"):
                ws = self.wb[name]
                for r in range(1, ws.max_row + 1):
                    val = ws.cell(r, 9).value or ""  # Notes column
                    if "From fragment" in val:
                        return  # Found at least one
        # Fragments may not always resolve depending on file structure
        # This is a soft check — pass if no fragments detected
