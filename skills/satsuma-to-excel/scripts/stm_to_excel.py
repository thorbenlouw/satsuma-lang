#!/usr/bin/env python3
"""Deterministic Satsuma-to-Excel converter.

Shells out to the satsuma CLI for all parsing, then generates a professional
Excel workbook using openpyxl.  No LLM involvement — same input always
produces the same output (excluding timestamp metadata).

Usage:
    stm_to_excel.py <file.stm> [extra.stm ...] -o output.xlsx [options]
    stm_to_excel.py examples/db-to-db.stm -o migration.xlsx
    stm_to_excel.py main.stm common.stm -o export.xlsx --targets postgres_db
"""

from __future__ import annotations

import argparse
import json
import os
import re
import subprocess
import sys
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ── Colour palette (PRD §Styling Specification) ────────────────────────

HEADER_BG = "2F3542"
HEADER_FG = "FFFFFF"
ALT_ROW_EVEN = "F8F9FA"
ALT_ROW_ODD = "FFFFFF"
WARNING_BG = "FFF3CD"
WARNING_FG = "856404"
QUESTION_BG = "CCE5FF"
QUESTION_FG = "004085"
COMPUTED_BG = "F0F0F0"
COMPUTED_FG = "6C757D"
PK_BG = "E8EDF2"
PII_FG = "6F42C1"
ARROW_FG = "ADB5BD"
SNAPSHOT_BG = WARNING_BG
SNAPSHOT_FG = WARNING_FG

# ── Fonts ───────────────────────────────────────────────────────────────

FONT_TITLE = Font(name="Calibri", size=16, bold=True, color="2F3542")
FONT_SECTION = Font(name="Calibri", size=12, bold=True)
FONT_META = Font(name="Calibri", size=10, color="999999")
FONT_META_ITALIC = Font(name="Calibri", size=10, color="999999", italic=True)
FONT_HEADER = Font(name="Calibri", size=10, bold=True, color=HEADER_FG)
FONT_DATA = Font(name="Calibri", size=10)
FONT_DATA_ITALIC_GRAY = Font(name="Calibri", size=10, italic=True, color=COMPUTED_FG)
FONT_SNAPSHOT = Font(name="Calibri", size=9, italic=True, color=SNAPSHOT_FG)
FONT_ARROW = Font(name="Calibri", size=10, color=ARROW_FG)
FONT_WARNING = Font(name="Calibri", size=10, color=WARNING_FG)
FONT_QUESTION = Font(name="Calibri", size=10, color=QUESTION_FG)
FONT_PII = Font(name="Calibri", size=10, color=PII_FG)
FONT_HYPERLINK = Font(name="Calibri", size=10, color="0563C1", underline="single")

# ── Fills ───────────────────────────────────────────────────────────────

FILL_HEADER = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
FILL_ALT_EVEN = PatternFill(start_color=ALT_ROW_EVEN, end_color=ALT_ROW_EVEN, fill_type="solid")
FILL_ALT_ODD = PatternFill(start_color=ALT_ROW_ODD, end_color=ALT_ROW_ODD, fill_type="solid")
FILL_WARNING = PatternFill(start_color=WARNING_BG, end_color=WARNING_BG, fill_type="solid")
FILL_QUESTION = PatternFill(start_color=QUESTION_BG, end_color=QUESTION_BG, fill_type="solid")
FILL_COMPUTED = PatternFill(start_color=COMPUTED_BG, end_color=COMPUTED_BG, fill_type="solid")
FILL_PK = PatternFill(start_color=PK_BG, end_color=PK_BG, fill_type="solid")
FILL_SNAPSHOT = PatternFill(start_color=SNAPSHOT_BG, end_color=SNAPSHOT_BG, fill_type="solid")
FILL_LIGHT_GRAY = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_WRAP = Alignment(wrap_text=True, vertical="top")
ALIGN_TOP = Alignment(vertical="top")

THIN_BORDER_BOTTOM = Border(bottom=Side(style="thin", color="DDDDDD"))


# ── Data model ──────────────────────────────────────────────────────────

@dataclass
class Comment:
    kind: str  # "warning" or "question"
    text: str
    location: str  # e.g. "Schema legacy_sqlserver"
    block: str = ""
    block_type: str = ""


@dataclass
class FieldMeta:
    name: str
    type: str = ""
    is_pk: bool = False
    is_required: bool = False
    is_unique: bool = False
    default: str | None = None
    tags: list[str] = field(default_factory=list)
    enum_values: list[str] = field(default_factory=list)
    note: str | None = None
    ref: str | None = None
    fragment_origin: str | None = None
    comments: list[Comment] = field(default_factory=list)


@dataclass
class SchemaInfo:
    name: str
    role: str  # "source" or "target"
    note: str | None = None
    fields: list[FieldMeta] = field(default_factory=list)


@dataclass
class MapEntry:
    """A single key-value in a map {} block, used for lookup tabs."""
    key: str
    value: str


@dataclass
class MappingArrow:
    source_field: str | None  # None for computed/derived
    source_type: str | None = None
    target_field: str = ""
    target_type: str | None = None
    transform_raw: str | None = None
    transform_human: str = ""
    classification: str = "none"
    is_derived: bool = False
    nl_text: str | None = None
    tags: list[str] = field(default_factory=list)
    is_required: bool = False
    notes: list[Comment] = field(default_factory=list)
    children: list[MappingArrow] = field(default_factory=list)
    map_entries: list[MapEntry] = field(default_factory=list)


@dataclass
class MappingInfo:
    name: str
    sources: list[str] = field(default_factory=list)
    targets: list[str] = field(default_factory=list)
    note: str | None = None
    arrows: list[MappingArrow] = field(default_factory=list)


@dataclass
class LookupTable:
    name: str
    entries: list[MapEntry] = field(default_factory=list)


@dataclass
class WorkbookData:
    title: str = ""
    note: str | None = None
    metadata: dict[str, str] = field(default_factory=dict)
    tags: list[str] = field(default_factory=list)
    source_files: list[str] = field(default_factory=list)
    schemas: list[SchemaInfo] = field(default_factory=list)
    mappings: list[MappingInfo] = field(default_factory=list)
    issues: list[Comment] = field(default_factory=list)
    lookups: list[LookupTable] = field(default_factory=list)
    timestamp: str = ""


# ── Transform translation ──────────────────────────────────────────────

# Ordered list of (pattern, replacement) for known pipeline tokens.
# Applied in order; first match wins per token.
TRANSFORM_TRANSLATIONS: list[tuple[str, str]] = [
    # Parameterised functions (must come before bare names)
    (r"^coalesce\((.+?)\)$", r"default to \1 if null"),
    (r"^round\((\d+)\)$", r"round to \1 decimal places"),
    (r"^round$", "round to nearest integer"),
    (r"^truncate\((\d+)\)$", r"truncate to \1 characters"),
    (r"^max_length\((\d+)\)$", r"limit to \1 characters"),
    (r"^pad_left\((\d+),\s*\"(.+?)\"\)$", r'pad left to \1 chars with "\2"'),
    (r"^pad_right\((\d+),\s*\"(.+?)\"\)$", r'pad right to \1 chars with "\2"'),
    (r"^prepend\(\"(.+?)\"\)$", r'prepend "\1"'),
    (r"^append\(\"(.+?)\"\)$", r'append "\1"'),
    (r"^replace\(\"(.+?)\",\s*\"(.+?)\"\)$", r'replace "\1" with "\2"'),
    (r"^split\(\"(.+?)\"\)$", r'split on "\1"'),
    (r"^uuid_v5\(.+?\)$", "generate UUID v5"),
    (r"^encrypt\((.+?),\s*.+?\)$", r"encrypt (\1)"),
    (r"^hash\((.+?)\)$", r"hash (\1)"),
    (r"^parse\(\"(.+?)\"\)$", r'parse as "\1"'),
    # Arithmetic
    (r"^\*\s*(\d+)$", r"multiply by \1"),
    (r"^/\s*(\d+)$", r"divide by \1"),
    (r"^\+\s*(\d+)$", r"add \1"),
    (r"^-\s*(\d+)$", r"subtract \1"),
    # Bare-name tokens
    (r"^trim$", "trim whitespace"),
    (r"^lowercase$", "convert to lowercase"),
    (r"^uppercase$", "convert to uppercase"),
    (r"^title_case$", "convert to title case"),
    (r"^null_if_empty$", "set null if empty"),
    (r"^null_if_invalid$", "set null if invalid"),
    (r"^validate_email$", "validate email format"),
    (r"^escape_html$", "escape HTML characters"),
    (r"^to_string$", "convert to text"),
    (r"^to_number$", "convert to number"),
    (r"^to_boolean$", "convert to boolean"),
    (r"^to_utc$", "convert to UTC"),
    (r"^to_iso8601$", "format as ISO 8601"),
    (r"^to_e164$", "format as E.164 phone"),
    (r"^now_utc\(\)$", "current UTC timestamp"),
    (r"^first$", "take first element"),
    (r"^last$", "take last element"),
    (r"^drop_if_invalid$", "drop if invalid"),
    (r"^error_if_null$", "error if null"),
    (r"^warn_if_invalid$", "warn if invalid"),
    (r"^assume_utc$", "assume UTC timezone"),
]


def _translate_single_token(token: str) -> str:
    """Translate one pipeline token to human-readable form."""
    token = token.strip()
    if not token:
        return ""
    # NL string — strip quotes, pass through verbatim
    if (token.startswith('"') and token.endswith('"')) or (
        token.startswith("'") and token.endswith("'")
    ):
        return token[1:-1]
    for pattern, replacement in TRANSFORM_TRANSLATIONS:
        m = re.match(pattern, token)
        if m:
            return m.expand(replacement)
    # Unknown token — return as-is
    return token


def _parse_map_block(raw: str) -> tuple[str, list[MapEntry], bool]:
    """Parse a map { ... } block.

    Returns (inline_text, entries, is_conditional).
    - Simple maps: inline_text = 'A = "active", S = "suspended"'
    - Conditional maps: inline_text = 'Conditional — see detail', entries populated
    """
    # Extract the content between map { ... }
    m = re.match(r"^map\s*\{(.*)\}$", raw.strip(), re.DOTALL)
    if not m:
        return raw, [], False

    body = m.group(1).strip()
    entries: list[MapEntry] = []
    is_conditional = False

    for line in body.split("\n"):
        line = line.strip()
        if not line:
            continue
        # Match key: value patterns
        kv_match = re.match(r'^([^:]+?):\s*(.+)$', line)
        if kv_match:
            key = kv_match.group(1).strip()
            value = kv_match.group(2).strip()
            # Strip trailing comma if present
            value = value.rstrip(",")
            # Check for conditional operators
            if re.match(r'^[<>=!]', key) or key == "default":
                is_conditional = True
            entries.append(MapEntry(key=key, value=value))

    if is_conditional:
        parts = []
        for e in entries:
            if e.key == "default":
                parts.append(f"default = {e.value}")
            else:
                parts.append(f"{e.key} = {e.value}")
        return "Conditional \u2014 see detail", entries, True
    else:
        # Simple map — inline as key = value pairs
        parts = []
        for e in entries:
            k = e.key
            if k == "null":
                k = "(empty)"
            elif k in ("_", "default"):
                k = "(other)"
            parts.append(f"{k} = {e.value}")
        return ", ".join(parts), entries, False


def translate_transform(
    transforms: list[str] | None,
    nl_text: str | None,
    classification: str,
) -> tuple[str, list[MapEntry], bool]:
    """Translate a Satsuma transform to human-readable text.

    Returns (human_text, map_entries_for_child_rows, is_conditional).
    """
    if classification == "nl" and nl_text:
        # Pure NL — strip quotes and pass through
        text = nl_text.strip()
        if text.startswith('"') and text.endswith('"'):
            text = text[1:-1]
        return text.strip(), [], False

    parts: list[str] = []
    all_entries: list[MapEntry] = []
    is_conditional = False

    if transforms:
        for t in transforms:
            t = t.strip()
            if not t:
                continue
            # map { ... } block
            if t.startswith("map"):
                inline, entries, cond = _parse_map_block(t)
                parts.append(inline)
                all_entries.extend(entries)
                if cond:
                    is_conditional = True
            else:
                parts.append(_translate_single_token(t))

    # If mixed classification, append the NL text
    if classification == "mixed" and nl_text:
        nl = nl_text.strip()
        if nl.startswith('"') and nl.endswith('"'):
            nl = nl[1:-1]
        parts.append(nl.strip())

    human = " \u2192 ".join(p for p in parts if p)
    return human, all_entries, is_conditional


# ── Data collection ─────────────────────────────────────────────────────

def _run_satsuma(args: list[str]) -> str:
    """Run a satsuma CLI command and return stdout.

    The CLI may write warnings to stderr and still return exit code 1
    (e.g. unresolved imports), so we accept any exit code as long as
    stdout contains usable output.  We only fail if there is *no*
    stdout at all and the exit code is non-zero.
    """
    cmd = ["npx", "satsuma"] + args
    result = subprocess.run(
        cmd,
        capture_output=True,
        text=True,
        timeout=30,
    )
    if result.returncode != 0 and not result.stdout.strip():
        print(f"Error running: {' '.join(cmd)}", file=sys.stderr)
        print(result.stderr, file=sys.stderr)
        sys.exit(1)
    return result.stdout


def _run_satsuma_json(args: list[str]) -> dict | list:
    """Run a satsuma CLI command and return parsed JSON."""
    raw = _run_satsuma(args)
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        print(f"Error: invalid JSON from satsuma {' '.join(args)}", file=sys.stderr)
        print(raw[:500], file=sys.stderr)
        sys.exit(1)


def collect_data(
    stm_files: list[str],
    targets: list[str] | None = None,
    title_override: str | None = None,
    timestamp: str | None = None,
) -> WorkbookData:
    """Collect all data from the satsuma CLI for workbook generation."""
    # Use the first file as the primary path for CLI commands
    primary = stm_files[0]
    # If multiple files, use the directory
    path = primary if len(stm_files) == 1 else str(Path(primary).parent)

    # 1. Graph — primary data source
    graph = _run_satsuma_json(["graph", "--json", path])

    # 2. Summary — for stats
    summary = _run_satsuma_json(["summary", "--json", path])

    # 3. NL — for integration note
    nl_items = _run_satsuma_json(["nl", "all", "--json", path])

    # 4. Warnings + Questions
    warnings_data = _run_satsuma_json(["warnings", "--json", path])
    questions_data = _run_satsuma_json(["warnings", "--questions", "--json", path])

    # -- Build integration info --
    integration_note = None
    for nl in nl_items:
        if nl.get("parent") is None and nl.get("kind") == "note":
            integration_note = nl["text"]
            break

    # Derive title
    title = title_override
    if not title and integration_note:
        # Extract first heading
        for line in integration_note.split("\n"):
            line = line.strip()
            if line.startswith("# "):
                title = line[2:].strip()
                break
    if not title:
        title = Path(primary).stem.replace("-", " ").replace("_", " ").title()

    # -- Build issues list --
    issues: list[Comment] = []
    for item in warnings_data.get("items", []):
        issues.append(Comment(
            kind="warning",
            text=item["text"],
            location=f"{item.get('blockType', 'block').title()} {item.get('block', '?')}",
            block=item.get("block", ""),
            block_type=item.get("blockType", ""),
        ))
    for item in questions_data.get("items", []):
        issues.append(Comment(
            kind="question",
            text=item["text"],
            location=f"{item.get('blockType', 'block').title()} {item.get('block', '?')}",
            block=item.get("block", ""),
            block_type=item.get("blockType", ""),
        ))

    # -- Build schema map from graph nodes --
    graph_schemas: dict[str, dict] = {}
    graph_fragments: dict[str, dict] = {}
    graph_mappings: list[dict] = []
    for node in graph.get("nodes", []):
        if node["kind"] == "schema":
            graph_schemas[node["id"]] = node
        elif node["kind"] == "fragment":
            graph_fragments[node["id"]] = node
        elif node["kind"] == "mapping":
            graph_mappings.append(node)

    # -- Determine schema roles from mappings --
    source_names: set[str] = set()
    target_names: set[str] = set()
    for gm in graph_mappings:
        for s in gm.get("sources", []):
            source_names.add(s)
        for t in gm.get("targets", []):
            target_names.add(t)

    # -- Target scoping --
    if targets:
        target_set = set(targets)
        # Filter mappings to those targeting requested schemas
        graph_mappings = [
            m for m in graph_mappings
            if any(t in target_set for t in m.get("targets", []))
        ]
        # Re-derive sources from filtered mappings
        source_names = set()
        target_names = set()
        for gm in graph_mappings:
            for s in gm.get("sources", []):
                source_names.add(s)
            for t in gm.get("targets", []):
                target_names.add(t)

    # -- Build SchemaInfo with full field metadata --
    schemas: list[SchemaInfo] = []
    all_schema_names = source_names | target_names
    for sname in all_schema_names:
        role = "target" if sname in target_names else "source"
        # Get field metadata from CLI
        try:
            fields_json = _run_satsuma_json(["fields", sname, "--json", path])
        except SystemExit:
            fields_json = []

        # Get schema note
        schema_note = None
        gs = graph_schemas.get(sname)
        if gs:
            schema_note = gs.get("note")

        # Parse field metadata
        fields: list[FieldMeta] = []
        for fj in fields_json:
            fm = FieldMeta(name=fj["name"], type=fj.get("type", ""))
            for md in fj.get("metadata", []):
                if md["kind"] == "tag":
                    tag = md["tag"]
                    if tag == "pk":
                        fm.is_pk = True
                    elif tag == "required":
                        fm.is_required = True
                    elif tag == "unique":
                        fm.is_unique = True
                    else:
                        fm.tags.append(tag)
                elif md["kind"] == "enum":
                    fm.enum_values = md.get("values", [])
                elif md["kind"] == "kv":
                    key = md["key"]
                    val = md.get("value", "")
                    if key == "default":
                        fm.default = val
                    elif key == "ref":
                        fm.ref = val
                    elif key == "note":
                        fm.note = val
                    else:
                        fm.tags.append(f"{key} {val}")
            fields.append(fm)

        # Detect fragment origins by comparing with fragment field lists
        if gs:
            graph_fields = {f["name"] for f in gs.get("fields", [])}
            for frag_name, frag_node in graph_fragments.items():
                frag_fields = {f["name"] for f in frag_node.get("fields", [])}
                if frag_fields and frag_fields <= graph_fields:
                    for fm in fields:
                        if fm.name in frag_fields:
                            fm.fragment_origin = frag_name

        schemas.append(SchemaInfo(
            name=sname,
            role=role,
            note=schema_note,
            fields=fields,
        ))

    # Sort: targets first (file order), then sources (file order)
    target_schemas = [s for s in schemas if s.role == "target"]
    source_schemas = [s for s in schemas if s.role == "source"]
    schemas = target_schemas + source_schemas

    # -- Build MappingInfo from graph edges --
    # Group edges by mapping name
    edges_by_mapping: dict[str, list[dict]] = {}
    for edge in graph.get("edges", []):
        mname = edge.get("mapping", "")
        edges_by_mapping.setdefault(mname, []).append(edge)

    # Build lookup tables from map {} blocks
    lookups: dict[str, LookupTable] = {}

    mappings: list[MappingInfo] = []
    for gm in graph_mappings:
        mname = gm["id"]
        # Get mapping note from NL
        mapping_note = None
        for nl in nl_items:
            if nl.get("parent") == mname and nl.get("kind") == "note":
                mapping_note = nl["text"]
                break

        arrows: list[MappingArrow] = []
        for edge in edges_by_mapping.get(mname, []):
            src_field = edge.get("from")
            tgt_field = edge.get("to")
            if src_field:
                # Strip schema prefix (e.g. "legacy_sqlserver.CUST_ID" -> "CUST_ID")
                src_field = src_field.split(".")[-1] if "." in src_field else src_field
            if tgt_field:
                tgt_field = tgt_field.split(".")[-1] if "." in tgt_field else tgt_field

            edge_transforms = edge.get("transforms", [])
            nl_text = edge.get("nl_text")
            classification = edge.get("classification", "none")
            is_derived = edge.get("derived", False)

            human, map_entries, is_conditional = translate_transform(
                edge_transforms, nl_text, classification,
            )

            # Look up source/target types from schemas
            src_type = None
            tgt_type = None
            tgt_required = False
            tgt_tags: list[str] = []
            src_tags: list[str] = []

            if src_field:
                for s in schemas:
                    if s.role == "source":
                        for f in s.fields:
                            if f.name == src_field:
                                src_type = f.type
                                src_tags = list(f.tags)
                                break
            if tgt_field:
                for s in schemas:
                    if s.role == "target":
                        for f in s.fields:
                            if f.name == tgt_field:
                                tgt_type = f.type
                                tgt_required = f.is_required
                                tgt_tags = list(f.tags)
                                break

            merged_tags = list(dict.fromkeys(src_tags + tgt_tags))

            # Extract map {} entries for lookup tab
            for t in edge_transforms:
                if t.strip().startswith("map"):
                    _, entries, _ = _parse_map_block(t.strip())
                    if entries:
                        lookup_name = f"{tgt_field}_map" if tgt_field else "lookup"
                        if lookup_name not in lookups:
                            lookups[lookup_name] = LookupTable(
                                name=lookup_name, entries=entries,
                            )

            arrow = MappingArrow(
                source_field=src_field if not is_derived else None,
                source_type=src_type,
                target_field=tgt_field or "",
                target_type=tgt_type,
                transform_raw="; ".join(edge_transforms) if edge_transforms else None,
                transform_human=human,
                classification=classification,
                is_derived=is_derived,
                nl_text=nl_text,
                tags=merged_tags,
                is_required=tgt_required,
                map_entries=map_entries if is_conditional else [],
            )
            arrows.append(arrow)

        mappings.append(MappingInfo(
            name=mname,
            sources=gm.get("sources", []),
            targets=gm.get("targets", []),
            note=mapping_note,
            arrows=arrows,
        ))

    ts = timestamp or datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    return WorkbookData(
        title=title,
        note=integration_note,
        metadata={},
        tags=[],
        source_files=[str(Path(f).name) for f in stm_files],
        schemas=schemas,
        mappings=mappings,
        issues=issues,
        lookups=list(lookups.values()),
        timestamp=ts,
    )


# ── Workbook generation ─────────────────────────────────────────────────

def _set_header_row(ws: Worksheet, row: int, headers: list[str]) -> None:
    """Write a styled header row."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER_BOTTOM


def _alt_fill(row_idx: int) -> PatternFill:
    """Return alternating row fill (1-indexed data row)."""
    return FILL_ALT_EVEN if row_idx % 2 == 0 else FILL_ALT_ODD


def _apply_data_tab_formatting(
    ws: Worksheet,
    header_row: int,
    data_end_row: int,
    col_count: int,
    freeze_col: int = 0,
) -> None:
    """Apply common formatting: freeze panes, auto-filter, print layout."""
    # Freeze panes
    if freeze_col > 0:
        ws.freeze_panes = ws.cell(row=header_row + 1, column=freeze_col + 1)
    else:
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # Auto-filter
    if data_end_row >= header_row:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(col_count)}{data_end_row}"

    # Print layout
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(
        fitToPage=True,
    )
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = f"{header_row}:{header_row}"


def _set_column_widths(ws: Worksheet, widths: dict[int, float]) -> None:
    """Set explicit column widths (1-indexed)."""
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def create_overview_tab(ws: Worksheet, data: WorkbookData) -> None:
    """Generate the Overview tab."""
    row = 1

    # Title
    ws.cell(row=row, column=1, value=data.title).font = FONT_TITLE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 2

    # Metadata line
    meta_parts = []
    if data.metadata.get("version"):
        meta_parts.append(f"Version: {data.metadata['version']}")
    if data.metadata.get("author"):
        meta_parts.append(f"Author: {data.metadata['author']}")
    if data.metadata.get("cardinality"):
        meta_parts.append(f"Cardinality: {data.metadata['cardinality']}")
    if meta_parts:
        ws.cell(row=row, column=1, value="  \u00b7  ".join(meta_parts)).font = FONT_META
        row += 1

    # Tags
    if data.tags:
        ws.cell(row=row, column=1, value=", ".join(data.tags)).font = FONT_META_ITALIC
        row += 1

    if meta_parts or data.tags:
        row += 1

    # Note content
    if data.note:
        # Strip the title heading if present
        note_text = data.note
        lines = note_text.split("\n")
        if lines and lines[0].strip().startswith("# "):
            note_text = "\n".join(lines[1:]).strip()
        if note_text:
            cell = ws.cell(row=row, column=1, value=note_text)
            cell.font = FONT_DATA
            cell.fill = FILL_LIGHT_GRAY
            cell.alignment = ALIGN_WRAP
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            # Estimate row height based on line count
            line_count = note_text.count("\n") + 1
            ws.row_dimensions[row].height = max(15, line_count * 15)
            row += 2

    # Systems table
    ws.cell(row=row, column=1, value="Systems").font = FONT_SECTION
    row += 1

    sys_headers = ["Schema", "Role", "Description"]
    for col, h in enumerate(sys_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
    row += 1

    for schema in data.schemas:
        ws.cell(row=row, column=1, value=schema.name).font = FONT_DATA
        ws.cell(row=row, column=2, value=schema.role.title()).font = FONT_DATA
        ws.cell(row=row, column=3, value=schema.note or "").font = FONT_DATA
        fill = _alt_fill(row)
        for c in range(1, 4):
            ws.cell(row=row, column=c).fill = fill
        row += 1

    row += 1

    # Table of Contents
    ws.cell(row=row, column=1, value="Contents").font = FONT_SECTION
    row += 1

    # We'll add hyperlinks to each tab
    toc_entries: list[tuple[str, str]] = []
    toc_entries.append(("Issues", "Warnings and open questions"))
    for m in data.mappings:
        src = ", ".join(m.sources) if m.sources else "?"
        tgt = ", ".join(m.targets) if m.targets else "?"
        tab_name = _mapping_tab_name(m)
        toc_entries.append((tab_name, f"Mapping: {src} \u2192 {tgt}"))

    for s in data.schemas:
        prefix = "Tgt" if s.role == "target" else "Src"
        tab_name = _schema_tab_name(s)
        toc_entries.append((tab_name, f"{s.role.title()} schema: {len(s.fields)} fields"))

    for lt in data.lookups:
        tab_name = _lookup_tab_name(lt)
        toc_entries.append((tab_name, f"Lookup: {len(lt.entries)} entries"))

    for tab_name, desc in toc_entries:
        cell = ws.cell(row=row, column=1, value=tab_name)
        cell.font = FONT_HYPERLINK
        # Internal hyperlink to sheet
        safe_name = tab_name.replace("'", "''")
        cell.hyperlink = f"#'{safe_name}'!A1"
        ws.cell(row=row, column=2, value=desc).font = FONT_DATA
        row += 1

    row += 1

    # Snapshot warning
    source_list = ", ".join(data.source_files)
    warning_text = (
        f"This is a read-only snapshot. The definitive source of truth is the "
        f".stm file(s). Generated from: {source_list} on {data.timestamp}"
    )
    cell = ws.cell(row=row, column=1, value=warning_text)
    cell.font = FONT_SNAPSHOT
    cell.fill = FILL_SNAPSHOT
    cell.alignment = ALIGN_WRAP
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15


def create_issues_tab(ws: Worksheet, data: WorkbookData) -> None:
    """Generate the Issues tab."""
    headers = ["#", "Location", "Type", "Description"]
    col_widths = {1: 5, 2: 30, 3: 12, 4: 80}
    _set_column_widths(ws, col_widths)

    _set_header_row(ws, 1, headers)

    if not data.issues:
        ws.merge_cells("A2:D2")
        ws.cell(row=2, column=1, value="No warnings or open questions found.").font = FONT_DATA
        _apply_data_tab_formatting(ws, 1, 2, 4)
        return

    for idx, issue in enumerate(data.issues, 1):
        row = idx + 1
        ws.cell(row=row, column=1, value=idx).font = FONT_DATA
        ws.cell(row=row, column=2, value=issue.location).font = FONT_DATA
        ws.cell(row=row, column=3, value=issue.kind.title()).font = FONT_DATA
        ws.cell(row=row, column=4, value=issue.text).font = FONT_DATA
        ws.cell(row=row, column=4).alignment = ALIGN_WRAP

        # Row styling by type
        if issue.kind == "warning":
            fill = FILL_WARNING
            font = FONT_WARNING
        else:
            fill = FILL_QUESTION
            font = FONT_QUESTION

        for c in range(1, 5):
            ws.cell(row=row, column=c).fill = fill
            ws.cell(row=row, column=c).font = font

    _apply_data_tab_formatting(ws, 1, len(data.issues) + 1, 4)


def _mapping_tab_name(m: MappingInfo) -> str:
    """Generate tab name for a mapping."""
    if m.sources and m.targets:
        src = m.sources[0]
        tgt = m.targets[0]
        name = f"Map - {src} to {tgt}"
    elif m.name:
        name = f"Map - {m.name}"
    else:
        name = "Map"
    # Excel tab names limited to 31 chars
    return name[:31]


def _schema_tab_name(s: SchemaInfo) -> str:
    """Generate tab name for a schema."""
    prefix = "Tgt" if s.role == "target" else "Src"
    name = f"{prefix} - {s.name}"
    return name[:31]


def _lookup_tab_name(lt: LookupTable) -> str:
    """Generate tab name for a lookup."""
    name = f"Ref - {lt.name}"
    return name[:31]


def create_mapping_tab(ws: Worksheet, mapping: MappingInfo, data: WorkbookData) -> None:
    """Generate a mapping tab."""
    headers = ["#", "Source", "Source Type", "", "Target", "Target Type",
               "Req", "Transform", "Tags", "Notes"]
    col_widths = {1: 5, 2: 25, 3: 15, 4: 3, 5: 25, 6: 15, 7: 5, 8: 50, 9: 15, 10: 40}
    _set_column_widths(ws, col_widths)

    start_row = 1

    # Mapping-level note (merged row above data)
    if mapping.note:
        cell = ws.cell(row=start_row, column=1, value=mapping.note)
        cell.font = FONT_DATA
        cell.fill = FILL_LIGHT_GRAY
        cell.alignment = ALIGN_WRAP
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=10)
        line_count = mapping.note.count("\n") + 1
        ws.row_dimensions[start_row].height = max(15, line_count * 15)
        start_row += 1

    header_row = start_row
    _set_header_row(ws, header_row, headers)
    # Arrow column header is just empty
    ws.cell(row=header_row, column=4, value="\u2192")

    row = header_row + 1
    for idx, arrow in enumerate(mapping.arrows, 1):
        # Source field
        if arrow.is_derived or arrow.source_field is None:
            ws.cell(row=row, column=2, value="\u2014").font = FONT_DATA_ITALIC_GRAY
            ws.cell(row=row, column=3, value="").font = FONT_DATA
        else:
            ws.cell(row=row, column=2, value=arrow.source_field).font = FONT_DATA
            ws.cell(row=row, column=3, value=arrow.source_type or "").font = FONT_DATA

        # Row number
        ws.cell(row=row, column=1, value=idx).font = FONT_DATA

        # Arrow
        ws.cell(row=row, column=4, value="\u2192").font = FONT_ARROW
        ws.cell(row=row, column=4).alignment = ALIGN_CENTER

        # Target
        ws.cell(row=row, column=5, value=arrow.target_field).font = FONT_DATA
        ws.cell(row=row, column=6, value=arrow.target_type or "").font = FONT_DATA

        # Required
        ws.cell(row=row, column=7, value="Yes" if arrow.is_required else "").font = FONT_DATA

        # Transform
        ws.cell(row=row, column=8, value=arrow.transform_human).font = FONT_DATA
        ws.cell(row=row, column=8).alignment = ALIGN_WRAP

        # Tags
        ws.cell(row=row, column=9, value=", ".join(arrow.tags) if arrow.tags else "").font = FONT_DATA

        # Notes
        note_text = ""
        if arrow.nl_text and arrow.classification == "mixed":
            # NL text already included in transform — add a note reference
            pass
        if arrow.notes:
            note_text = "; ".join(n.text for n in arrow.notes)
        ws.cell(row=row, column=10, value=note_text).font = FONT_DATA
        ws.cell(row=row, column=10).alignment = ALIGN_WRAP

        # Row fill
        if arrow.is_derived or arrow.source_field is None:
            fill = FILL_COMPUTED
            for c in range(1, 11):
                ws.cell(row=row, column=c).fill = fill
        else:
            fill = _alt_fill(idx)
            for c in range(1, 11):
                ws.cell(row=row, column=c).fill = fill

        row += 1

        # Child rows for conditional map blocks
        if arrow.map_entries:
            for ci, entry in enumerate(arrow.map_entries):
                child_label = f"{idx}{chr(97 + ci)}"  # 5a, 5b, 5c...
                ws.cell(row=row, column=1, value=child_label).font = FONT_DATA_ITALIC_GRAY

                if entry.key == "default":
                    display = f"default = {entry.value}"
                else:
                    display = f"{entry.key} = {entry.value}"
                ws.cell(row=row, column=8, value=display).font = FONT_DATA

                # Group this row
                ws.row_dimensions[row].outlineLevel = 1
                ws.row_dimensions[row].hidden = True

                for c in range(1, 11):
                    ws.cell(row=row, column=c).fill = _alt_fill(idx)

                row += 1

    _apply_data_tab_formatting(ws, header_row, row - 1, 10, freeze_col=1)


def create_schema_tab(ws: Worksheet, schema: SchemaInfo) -> None:
    """Generate a schema reference tab."""
    headers = ["#", "Field", "Type", "PK", "Required", "Unique",
               "Default", "Tags", "Notes"]
    col_widths = {1: 5, 2: 25, 3: 20, 4: 5, 5: 8, 6: 8, 7: 15, 8: 20, 9: 40}
    _set_column_widths(ws, col_widths)

    start_row = 1

    # Schema-level note
    if schema.note:
        cell = ws.cell(row=start_row, column=1, value=schema.note)
        cell.font = FONT_DATA
        cell.fill = FILL_LIGHT_GRAY
        cell.alignment = ALIGN_WRAP
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=9)
        start_row += 1

    header_row = start_row
    _set_header_row(ws, header_row, headers)

    # Track fragments for grouping
    current_fragment: str | None = None
    fragment_start_row: int | None = None

    row = header_row + 1
    for idx, f in enumerate(schema.fields, 1):
        # Fragment grouping: if this field has a fragment origin different from previous
        if f.fragment_origin and f.fragment_origin != current_fragment:
            current_fragment = f.fragment_origin
            fragment_start_row = row

        ws.cell(row=row, column=1, value=idx).font = FONT_DATA
        ws.cell(row=row, column=2, value=f.name).font = FONT_DATA
        ws.cell(row=row, column=3, value=f.type).font = FONT_DATA
        ws.cell(row=row, column=4, value="Yes" if f.is_pk else "").font = FONT_DATA
        ws.cell(row=row, column=5, value="Yes" if f.is_required else "").font = FONT_DATA
        ws.cell(row=row, column=6, value="Yes" if f.is_unique else "").font = FONT_DATA
        ws.cell(row=row, column=7, value=f.default or "").font = FONT_DATA

        # Tags — PII gets purple
        tags_text = ", ".join(f.tags) if f.tags else ""
        if f.enum_values:
            enum_str = "enum {" + ", ".join(f.enum_values) + "}"
            tags_text = f"{tags_text}, {enum_str}" if tags_text else enum_str
        if f.ref:
            ref_str = f"ref {f.ref}"
            tags_text = f"{tags_text}, {ref_str}" if tags_text else ref_str

        tags_cell = ws.cell(row=row, column=8, value=tags_text)
        if "pii" in f.tags:
            tags_cell.font = FONT_PII
        else:
            tags_cell.font = FONT_DATA

        # Notes
        note_parts = []
        if f.note:
            note_parts.append(f.note)
        if f.fragment_origin:
            note_parts.append(f"From fragment: {f.fragment_origin}")
        ws.cell(row=row, column=9, value="; ".join(note_parts) if note_parts else "").font = FONT_DATA
        ws.cell(row=row, column=9).alignment = ALIGN_WRAP

        # Row fill
        if f.is_pk:
            fill = FILL_PK
        else:
            fill = _alt_fill(idx)
        for c in range(1, 10):
            ws.cell(row=row, column=c).fill = fill

        # Fragment grouping
        if f.fragment_origin:
            ws.row_dimensions[row].outlineLevel = 1

        if not f.fragment_origin and current_fragment:
            current_fragment = None
            fragment_start_row = None

        row += 1

    _apply_data_tab_formatting(ws, header_row, row - 1, 9)


def create_lookup_tab(ws: Worksheet, lookup: LookupTable) -> None:
    """Generate a lookup/reference tab."""
    headers = ["Code", "Description"]
    col_widths = {1: 20, 2: 40}
    _set_column_widths(ws, col_widths)

    _set_header_row(ws, 1, headers)

    for idx, entry in enumerate(lookup.entries, 1):
        row = idx + 1
        ws.cell(row=row, column=1, value=entry.key).font = FONT_DATA
        ws.cell(row=row, column=2, value=entry.value).font = FONT_DATA
        fill = _alt_fill(idx)
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=2).fill = fill

    _apply_data_tab_formatting(ws, 1, len(lookup.entries) + 1, 2)


def generate_workbook(data: WorkbookData, output_path: str, options: dict) -> None:
    """Generate the complete Excel workbook."""
    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Tab 1: Overview
    ws_overview = wb.create_sheet("Overview")
    create_overview_tab(ws_overview, data)

    # Tab 2: Issues (unless --no-issues)
    if not options.get("no_issues"):
        ws_issues = wb.create_sheet("Issues")
        create_issues_tab(ws_issues, data)

    # Tab 3+: Mapping tabs
    for mapping in data.mappings:
        tab_name = _mapping_tab_name(mapping)
        ws_map = wb.create_sheet(tab_name)
        create_mapping_tab(ws_map, mapping, data)

    # Schema tabs (unless --no-schemas)
    if not options.get("no_schemas"):
        # Targets first, then sources (per PRD)
        for schema in data.schemas:
            if schema.role == "target":
                tab_name = _schema_tab_name(schema)
                ws_schema = wb.create_sheet(tab_name)
                create_schema_tab(ws_schema, schema)
        for schema in data.schemas:
            if schema.role == "source":
                tab_name = _schema_tab_name(schema)
                ws_schema = wb.create_sheet(tab_name)
                create_schema_tab(ws_schema, schema)

    # Lookup tabs (alphabetical)
    for lookup in sorted(data.lookups, key=lambda lt: lt.name):
        tab_name = _lookup_tab_name(lookup)
        ws_lookup = wb.create_sheet(tab_name)
        create_lookup_tab(ws_lookup, lookup)

    wb.save(output_path)


# ── CLI ──────────────────────────────────────────────────────────────────

def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="stm_to_excel",
        description="Convert Satsuma .stm files to professional Excel workbooks.",
    )
    parser.add_argument("stm_files", nargs="+", help="Input .stm file(s)")
    parser.add_argument("-o", "--output", required=True, help="Output .xlsx path")
    parser.add_argument("--targets", help="Comma-separated target schema names to include")
    parser.add_argument("--title", help="Override workbook title")
    parser.add_argument("--timestamp", help="Override generation timestamp (ISO 8601)")
    parser.add_argument("--no-issues", action="store_true", help="Omit the Issues tab")
    parser.add_argument("--no-schemas", action="store_true", help="Omit schema reference tabs")
    return parser


def main() -> None:
    parser = build_arg_parser()
    args = parser.parse_args()

    # Validate inputs
    for f in args.stm_files:
        if not Path(f).exists():
            print(f"Error: file not found: {f}", file=sys.stderr)
            sys.exit(1)
        if not f.endswith(".stm"):
            print(f"Error: expected .stm file, got: {f}", file=sys.stderr)
            sys.exit(1)

    targets = [t.strip() for t in args.targets.split(",")] if args.targets else None

    # Collect data
    data = collect_data(
        stm_files=args.stm_files,
        targets=targets,
        title_override=args.title,
        timestamp=args.timestamp,
    )

    # Generate workbook
    options = {
        "no_issues": args.no_issues,
        "no_schemas": args.no_schemas,
    }
    generate_workbook(data, args.output, options)

    # Report
    tab_count = 1  # Overview
    if not args.no_issues:
        tab_count += 1
    tab_count += len(data.mappings)
    if not args.no_schemas:
        tab_count += len(data.schemas)
    tab_count += len(data.lookups)

    print(f"Generated: {args.output}")
    print(f"  Tabs: {tab_count}")
    print(f"  Mappings: {len(data.mappings)}")
    print(f"  Schemas: {len(data.schemas)}")
    print(f"  Issues: {len(data.issues)}")
    if data.lookups:
        print(f"  Lookups: {len(data.lookups)}")


if __name__ == "__main__":
    main()
