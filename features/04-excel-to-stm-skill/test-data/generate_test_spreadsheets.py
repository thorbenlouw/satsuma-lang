#!/usr/bin/env python3
"""Generate diverse test spreadsheets for Excel-to-STM lite prompt testing."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


def style_header_row(ws, ncols, fill_color="1F4E79", font_color="FFFFFF"):
    """Apply consistent header styling."""
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color=font_color, size=11)
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def auto_width(ws):
    """Rough auto-fit column widths."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


# ──────────────────────────────────────────────────────────────
# Test 1: Simple 1:1 customer sync (single tab, clean, 12 rows)
# ──────────────────────────────────────────────────────────────

def create_simple_customer_sync():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Field Mapping"

    headers = [
        "Source System", "Source Field", "Source Data Type",
        "Target System", "Target Field", "Target Data Type",
        "Transformation Rule", "Required?", "Notes",
    ]
    ws.append(headers)
    style_header_row(ws, len(headers))

    rows = [
        ("SAP CRM", "KUNNR", "CHAR(10)", "Snowflake", "customer_id", "VARCHAR(10)", "Direct copy", "Y", "Primary key"),
        ("SAP CRM", "NAME1", "CHAR(35)", "Snowflake", "customer_name", "VARCHAR(100)", "Trim whitespace", "Y", ""),
        ("SAP CRM", "NAME2", "CHAR(35)", "Snowflake", "customer_name_2", "VARCHAR(100)", "Trim whitespace", "N", "Secondary name line"),
        ("SAP CRM", "STRAS", "CHAR(35)", "Snowflake", "street_address", "VARCHAR(200)", "Trim", "N", ""),
        ("SAP CRM", "ORT01", "CHAR(35)", "Snowflake", "city", "VARCHAR(100)", "Trim", "N", ""),
        ("SAP CRM", "PSTLZ", "CHAR(10)", "Snowflake", "postal_code", "VARCHAR(20)", "Trim, remove leading zeros for US", "N", ""),
        ("SAP CRM", "LAND1", "CHAR(3)", "Snowflake", "country_code", "VARCHAR(3)", "Direct, validate ISO-3166", "Y", ""),
        ("SAP CRM", "SMTP_ADDR", "CHAR(241)", "Snowflake", "email", "VARCHAR(255)", "Lowercase, validate email format", "N", "PII field"),
        ("SAP CRM", "TELF1", "CHAR(16)", "Snowflake", "phone_primary", "VARCHAR(20)", "Format to E.164", "N", "PII field"),
        ("SAP CRM", "ERDAT", "DATS(8)", "Snowflake", "created_date", "DATE", "Convert YYYYMMDD to ISO date", "Y", "SAP date format"),
        ("SAP CRM", "AEDAT", "DATS(8)", "Snowflake", "modified_date", "DATE", "Convert YYYYMMDD to ISO date", "N", ""),
        ("SAP CRM", "LOEVM", "CHAR(1)", "Snowflake", "is_deleted", "BOOLEAN", "X=true, space=false", "Y", "Deletion flag"),
    ]
    for r in rows:
        ws.append(r)

    # Highlight PII rows
    yellow = PatternFill("solid", fgColor="FFFFCC")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[8].value and "PII" in str(row[8].value):
            for cell in row:
                cell.fill = yellow

    auto_width(ws)
    path = os.path.join(SCRIPT_DIR, "test-01-simple-customer-sync.xlsx")
    wb.save(path)
    print(f"Created: {path}")


# ──────────────────────────────────────────────────────────────
# Test 2: Multi-tab order pipeline (mapping + lookups + changelog)
# ──────────────────────────────────────────────────────────────

def create_multi_tab_order_pipeline():
    wb = openpyxl.Workbook()

    # --- Tab 1: Instructions ---
    ws_instr = wb.active
    ws_instr.title = "Instructions"
    ws_instr["A1"] = "Order Migration Mapping — v2.3"
    ws_instr["A1"].font = Font(bold=True, size=16)
    ws_instr["A3"] = "This workbook documents the field-level mapping from the legacy Order Management System (OMS) to the new Order Service microservice."
    ws_instr["A5"] = "Tabs:"
    ws_instr["A6"] = "  • Order Header Mapping — main order fields"
    ws_instr["A7"] = "  • Order Line Mapping — line item fields"
    ws_instr["A8"] = "  • Status Codes — reference table for order/line statuses"
    ws_instr["A9"] = "  • Priority Codes — reference table for priority levels"
    ws_instr["A10"] = "  • Changelog — revision history"
    ws_instr["A12"] = "Colour coding: Yellow = needs review, Red = blocked, Strikethrough = deprecated"
    ws_instr.column_dimensions["A"].width = 80

    # --- Tab 2: Order Header Mapping ---
    ws_hdr = wb.create_sheet("Order Header Mapping")
    hdr_headers = [
        "Source Field (OMS)", "Type", "Target Field (Order Service)", "Type",
        "Transform", "Nullable?", "Status", "Comments",
    ]
    ws_hdr.append(hdr_headers)
    style_header_row(ws_hdr, len(hdr_headers), fill_color="2E75B6")

    hdr_rows = [
        ("ORDER_ID", "NUMBER(12)", "order_id", "UUID", "Generate UUID v5 from ORDER_ID using namespace 'orders'", "N", "Approved", ""),
        ("CUST_ID", "NUMBER(10)", "customer_ref", "STRING", "Prepend 'CUST-'", "N", "Approved", ""),
        ("ORDER_DATE", "DATE", "placed_at", "TIMESTAMP", "Convert to UTC timestamp, assume EST timezone", "N", "Approved", "All legacy dates are EST"),
        ("SHIP_DATE", "DATE", "shipped_at", "TIMESTAMP", "Convert to UTC, null if order not yet shipped", "Y", "Approved", ""),
        ("ORDER_STATUS", "CHAR(2)", "status", "STRING", "Map using Status Codes tab", "N", "Approved", "See lookup tab"),
        ("PRIORITY", "CHAR(1)", "priority", "STRING", "Map using Priority Codes tab", "Y", "Approved", "Null = 'standard'"),
        ("TOTAL_AMT", "NUMBER(12,2)", "total_amount", "DECIMAL(14,2)", "Direct copy", "N", "Approved", ""),
        ("CURRENCY", "CHAR(3)", "currency_code", "STRING(3)", "Uppercase, validate ISO-4217", "N", "Approved", ""),
        ("DISCOUNT_PCT", "NUMBER(5,2)", "discount_percent", "DECIMAL(5,2)", "Divide by 100 if > 1 (legacy stores as whole number)", "Y", "Needs Review", "Inconsistent format in source"),
        ("SALES_REP", "VARCHAR(50)", "assigned_rep", "STRING", "Trim", "Y", "Approved", ""),
        ("CHANNEL", "CHAR(3)", "sales_channel", "STRING", "WEB=online, POS=in_store, PHN=phone, EDI=wholesale", "N", "Approved", ""),
        ("NOTES", "CLOB", "internal_notes", "TEXT", "Truncate to 2000 chars, escape HTML entities", "Y", "Approved", "Legacy has no length limit"),
        ("CREATED_BY", "VARCHAR(30)", "created_by", "STRING", "Direct copy", "Y", "Deprecated", "No longer tracked in new system"),
        ("LEGACY_FLAG", "CHAR(1)", "—", "—", "Do not migrate", "—", "Blocked", "Field has no equivalent"),
    ]
    for r in hdr_rows:
        ws_hdr.append(r)

    # Style: yellow for "Needs Review", red for "Blocked", strikethrough for "Deprecated"
    yellow = PatternFill("solid", fgColor="FFFFCC")
    red = PatternFill("solid", fgColor="FFCCCC")
    for row in ws_hdr.iter_rows(min_row=2, max_row=ws_hdr.max_row):
        status = str(row[6].value) if row[6].value else ""
        if status == "Needs Review":
            for cell in row:
                cell.fill = yellow
        elif status == "Blocked":
            for cell in row:
                cell.fill = red
        elif status == "Deprecated":
            for cell in row:
                cell.font = Font(strikethrough=True, color="999999")

    auto_width(ws_hdr)

    # --- Tab 3: Order Line Mapping ---
    ws_line = wb.create_sheet("Order Line Mapping")
    line_headers = [
        "Source Field (OMS)", "Type", "Target Field (Order Service)", "Type",
        "Transform", "Nullable?", "Comments",
    ]
    ws_line.append(line_headers)
    style_header_row(ws_line, len(line_headers), fill_color="2E75B6")

    line_rows = [
        ("ORDER_ID", "NUMBER(12)", "order_id", "UUID", "Same UUID v5 transform as header", "N", "FK to order header"),
        ("LINE_NO", "NUMBER(5)", "line_number", "INTEGER", "Direct copy", "N", ""),
        ("PRODUCT_CODE", "VARCHAR(20)", "product_sku", "STRING", "Uppercase, trim", "N", ""),
        ("PRODUCT_DESC", "VARCHAR(200)", "product_name", "STRING", "Trim", "Y", ""),
        ("QTY", "NUMBER(8)", "quantity", "INTEGER", "Direct copy, must be > 0", "N", ""),
        ("UNIT_PRICE", "NUMBER(10,2)", "unit_price", "DECIMAL(12,2)", "Direct copy", "N", ""),
        ("LINE_TOTAL", "NUMBER(12,2)", "line_total", "DECIMAL(14,2)", "Recalculate as QTY * UNIT_PRICE (don't trust source)", "N", "Source data has rounding errors"),
        ("LINE_STATUS", "CHAR(2)", "status", "STRING", "Map using Status Codes tab (same as header)", "N", ""),
        ("WAREHOUSE", "CHAR(4)", "fulfillment_center", "STRING", "Trim", "Y", ""),
        ("TAX_RATE", "NUMBER(5,4)", "tax_rate", "DECIMAL(5,4)", "Direct copy", "Y", "Stored as decimal, e.g. 0.0825"),
    ]
    for r in line_rows:
        ws_line.append(r)
    auto_width(ws_line)

    # --- Tab 4: Status Codes (lookup) ---
    ws_status = wb.create_sheet("Status Codes")
    ws_status.append(["Code", "Legacy Description", "New Value", "Active?"])
    style_header_row(ws_status, 4, fill_color="548235")
    status_rows = [
        ("OP", "Open", "open", "Y"),
        ("IP", "In Progress", "processing", "Y"),
        ("SH", "Shipped", "shipped", "Y"),
        ("DL", "Delivered", "delivered", "Y"),
        ("CN", "Cancelled", "cancelled", "Y"),
        ("RT", "Returned", "returned", "Y"),
        ("HD", "On Hold", "on_hold", "Y"),
        ("BO", "Backordered", "backordered", "Y"),
        ("PF", "Partially Fulfilled", "partial", "Y"),
        ("XX", "Unknown/Legacy", "unknown", "N"),
    ]
    for r in status_rows:
        ws_status.append(r)
    auto_width(ws_status)

    # --- Tab 5: Priority Codes (lookup) ---
    ws_prio = wb.create_sheet("Priority Codes")
    ws_prio.append(["Code", "Description", "New Value"])
    style_header_row(ws_prio, 3, fill_color="548235")
    prio_rows = [
        ("H", "High / Rush", "rush"),
        ("N", "Normal", "standard"),
        ("L", "Low / Backfill", "economy"),
    ]
    for r in prio_rows:
        ws_prio.append(r)
    auto_width(ws_prio)

    # --- Tab 6: Changelog ---
    ws_log = wb.create_sheet("Changelog")
    ws_log.append(["Date", "Author", "Change"])
    style_header_row(ws_log, 3, fill_color="808080")
    log_rows = [
        ("2025-01-15", "J. Smith", "Initial draft"),
        ("2025-02-03", "A. Patel", "Added order line mapping tab"),
        ("2025-02-20", "J. Smith", "Added status and priority code lookups"),
        ("2025-03-01", "M. Chen", "Marked CREATED_BY as deprecated, added LEGACY_FLAG as blocked"),
        ("2025-03-10", "A. Patel", "Updated DISCOUNT_PCT notes — source format is inconsistent"),
    ]
    for r in log_rows:
        ws_log.append(r)
    auto_width(ws_log)

    path = os.path.join(SCRIPT_DIR, "test-02-multi-tab-order-pipeline.xlsx")
    wb.save(path)
    print(f"Created: {path}")


# ──────────────────────────────────────────────────────────────
# Test 3: Healthcare HL7-to-FHIR — messy, prose-heavy transforms,
# conditional logic, nested structures, non-obvious column layout
# ──────────────────────────────────────────────────────────────

def create_healthcare_hl7_to_fhir():
    wb = openpyxl.Workbook()

    # --- Tab 1: Patient Demographics ---
    ws = wb.active
    ws.title = "Patient Demographics"

    # Non-standard header layout: headers in row 2, row 1 is a title
    ws["A1"] = "HL7 ADT → FHIR Patient Resource"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:H1")

    headers = [
        "HL7 Segment.Field", "HL7 Component", "FHIR Path",
        "FHIR Type", "Cardinality", "Mapping Logic", "Conditions", "Open Questions",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")

    rows = [
        ("PID-3", "CX.1 (ID Number)", "Patient.identifier[0].value", "string", "1..1",
         "Direct copy", "", ""),
        ("PID-3", "CX.4 (Assigning Authority)", "Patient.identifier[0].system", "uri", "1..1",
         "Map authority code to URI: MRN → 'urn:oid:2.16.840.1.113883.1.13', SSN → 'http://hl7.org/fhir/sid/us-ssn'",
         "Only when CX.5 = 'MR' or 'SS'", "What if authority code is not MRN or SSN?"),
        ("PID-5", "XPN.1 (Family Name)", "Patient.name[0].family", "string", "1..1",
         "Trim, title case", "", ""),
        ("PID-5", "XPN.2 (Given Name)", "Patient.name[0].given[0]", "string", "0..*",
         "Trim, title case. If multiple given names separated by space, split into array elements.",
         "", ""),
        ("PID-5", "XPN.3 (Middle Name)", "Patient.name[0].given[1]", "string", "0..1",
         "Trim, title case. Append to given[] array.", "", ""),
        ("PID-5", "XPN.5 (Prefix)", "Patient.name[0].prefix[0]", "string", "0..1",
         "Direct copy", "Only if present", ""),
        ("PID-5", "XPN.7 (Name Type Code)", "Patient.name[0].use", "code", "0..1",
         "L=official, M=maiden, A=anonymous, N=nickname, S=pseudonym",
         "", "Do we need to handle type D (adopted)?"),
        ("PID-7", "(full field)", "Patient.birthDate", "date", "0..1",
         "Parse HL7 date format (YYYYMMDD or YYYYMM or YYYY) to FHIR date (YYYY-MM-DD). Partial dates: keep precision as-is.",
         "", ""),
        ("PID-8", "(full field)", "Patient.gender", "code", "0..1",
         "M=male, F=female, O=other, U=unknown, A=other, N=unknown",
         "If empty, omit field entirely (don't send null)", "Confirm: A→other and N→unknown correct per local policy?"),
        ("PID-11", "XAD.1 (Street)", "Patient.address[0].line[0]", "string", "0..*",
         "Direct copy", "", ""),
        ("PID-11", "XAD.2 (Other)", "Patient.address[0].line[1]", "string", "0..1",
         "Direct copy, only if non-empty", "Skip if blank", ""),
        ("PID-11", "XAD.3 (City)", "Patient.address[0].city", "string", "0..1",
         "Trim, title case", "", ""),
        ("PID-11", "XAD.4 (State)", "Patient.address[0].state", "string", "0..1",
         "Validate against US state codes. If invalid, pass through with warning.", "", ""),
        ("PID-11", "XAD.5 (Zip)", "Patient.address[0].postalCode", "string", "0..1",
         "Trim. If 9 digits, format as XXXXX-XXXX.", "", ""),
        ("PID-11", "XAD.6 (Country)", "Patient.address[0].country", "string", "0..1",
         "Map HL7 3-letter to ISO 3166-1 alpha-2 (USA→US, CAN→CA, MEX→MX, etc.)",
         "", "Need full mapping table for all country codes"),
        ("PID-11", "XAD.7 (Address Type)", "Patient.address[0].use", "code", "0..1",
         "H=home, B=work, M=temp (mailing), C=temp",
         "Only map known types; omit use if type not recognized", ""),
        ("PID-13", "XTN.1 (Telephone)", "Patient.telecom[0].value", "string", "0..*",
         "Strip non-numeric except leading +. Format to E.164 if possible. If not parseable, store as-is with //! warning.",
         "PRN (primary residence) or ORN (other residence) in XTN.2", ""),
        ("PID-13", "XTN.2 (Use Code)", "Patient.telecom[0].use", "code", "0..1",
         "PRN=home, WPN=work, ORN=home, BPN=mobile",
         "", "BPN→mobile is our local convention, not HL7 standard"),
        ("PID-13", "XTN.3 (Equipment Type)", "Patient.telecom[0].system", "code", "0..1",
         "PH=phone, FX=fax, CP=phone (mobile), Internet=email, BP=pager",
         "If CP, also set use=mobile", "Confirm pager handling — do we even need it?"),
        ("PID-19", "(full field)", "Patient.identifier[1].value", "string", "0..1",
         "Direct copy. System = 'http://hl7.org/fhir/sid/us-ssn'",
         "Only if SSN is present and patient consent flag (PD1-12) = 'Y'",
         "PII — must be encrypted at rest"),
        ("", "", "Patient.meta.lastUpdated", "instant", "1..1",
         "Set to current UTC timestamp at time of conversion", "", "Computed field, no HL7 source"),
    ]
    for r in rows:
        ws.append(r)

    # Orange fill for rows with open questions
    orange = PatternFill("solid", fgColor="FFF2CC")
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        if row[7].value:  # "Open Questions" column
            for cell in row:
                cell.fill = orange

    auto_width(ws)

    # --- Tab 2: Sample Data ---
    ws_sample = wb.create_sheet("Sample Data")
    ws_sample.append(["NOTE: These are synthetic test records, not real patient data"])
    ws_sample["A1"].font = Font(bold=True, color="FF0000", size=12)
    ws_sample.append([])
    ws_sample.append([
        "PID-3.1", "PID-3.4", "PID-3.5", "PID-5.1", "PID-5.2",
        "PID-7", "PID-8", "PID-11.1", "PID-11.3", "PID-11.4", "PID-11.5",
    ])
    sample_rows = [
        ("MRN001234", "HOSP_A", "MR", "DOE", "JOHN",
         "19850315", "M", "123 Main St", "Springfield", "IL", "62704"),
        ("SSN555667788", "SSA", "SS", "SMITH", "JANE MARIE",
         "199201", "F", "456 Oak Ave Apt 2B", "Chicago", "IL", "606011234"),
        ("MRN005678", "HOSP_A", "MR", "O'BRIEN", "SEAN",
         "1970", "M", "", "Boston", "MA", "02101"),
    ]
    for r in sample_rows:
        ws_sample.append(r)
    auto_width(ws_sample)

    path = os.path.join(SCRIPT_DIR, "test-03-healthcare-hl7-to-fhir.xlsx")
    wb.save(path)
    print(f"Created: {path}")


if __name__ == "__main__":
    create_simple_customer_sync()
    create_multi_tab_order_pipeline()
    create_healthcare_hl7_to_fhir()
    print("\nDone — 3 test spreadsheets created.")
