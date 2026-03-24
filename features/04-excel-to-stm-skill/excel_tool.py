#!/usr/bin/env python3
"""Excel interrogation tool for the Excel-to-Satsuma skill.

A stateless CLI with five subcommands that extract bounded slices of
information from .xlsx files and return structured Markdown.

Usage:
    excel_tool.py survey    <file.xlsx>
    excel_tool.py headers   <file.xlsx> <tab-name>
    excel_tool.py formatting <file.xlsx> <tab-name>
    excel_tool.py range     <file.xlsx> <tab-name> [--rows START:END] [--cols A:H]
    excel_tool.py lookup    <file.xlsx> <tab-name> [--max-rows 500]
"""

from __future__ import annotations

import argparse
import sys
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

# Maximum output size in characters before truncation.
MAX_OUTPUT_CHARS = 100_000


# ── Helpers ──────────────────────────────────────────────────────────

def open_workbook(path: str) -> openpyxl.Workbook:
    p = Path(path)
    if not p.exists():
        print(f"Error: file not found: {path}", file=sys.stderr)
        sys.exit(1)
    if p.suffix.lower() != ".xlsx":
        print(f"Error: expected .xlsx file, got {p.suffix}", file=sys.stderr)
        sys.exit(1)
    try:
        return openpyxl.load_workbook(path, read_only=False, data_only=True)
    except Exception as e:
        print(f"Error: cannot open workbook: {e}", file=sys.stderr)
        sys.exit(1)


def get_sheet(wb: openpyxl.Workbook, tab_name: str) -> Worksheet:
    if tab_name not in wb.sheetnames:
        print(
            f"Error: tab '{tab_name}' not found. "
            f"Available tabs: {', '.join(wb.sheetnames)}",
            file=sys.stderr,
        )
        sys.exit(1)
    return wb[tab_name]


def cell_value(cell: Cell) -> str:
    """Return a string representation of a cell value."""
    if cell.value is None:
        return ""
    return str(cell.value)


def md_table(headers: list[str], rows: list[list[str]]) -> str:
    """Render a Markdown table."""
    if not headers:
        return ""
    lines = ["| " + " | ".join(headers) + " |"]
    lines.append("| " + " | ".join("---" for _ in headers) + " |")
    for row in rows:
        padded = row + [""] * (len(headers) - len(row))
        escaped = [v.replace("|", "\\|").replace("\n", " ") for v in padded]
        lines.append("| " + " | ".join(escaped) + " |")
    return "\n".join(lines)


def effective_row_count(ws: Worksheet) -> int:
    """Count rows excluding trailing empty rows."""
    max_row = ws.max_row or 0
    for r in range(max_row, 0, -1):
        for c in range(1, (ws.max_column or 0) + 1):
            if ws.cell(row=r, column=c).value is not None:
                return r
        # entire row is empty, keep looking
    return 0


def effective_col_count(ws: Worksheet) -> int:
    """Count columns excluding trailing empty columns."""
    max_col = ws.max_column or 0
    for c in range(max_col, 0, -1):
        for r in range(1, (ws.max_row or 0) + 1):
            if ws.cell(row=r, column=c).value is not None:
                return c
    return 0


def truncate_output(text: str) -> str:
    """Truncate output if it exceeds MAX_OUTPUT_CHARS."""
    if len(text) <= MAX_OUTPUT_CHARS:
        return text
    truncated = text[:MAX_OUTPUT_CHARS]
    return truncated + "\n\n**WARNING**: Output truncated at {:,} characters.\n".format(
        MAX_OUTPUT_CHARS
    )


# ── Subcommands ──────────────────────────────────────────────────────

def cmd_survey(args: argparse.Namespace) -> str:
    """Survey workbook structure: tabs, row/col counts, previews."""
    wb = open_workbook(args.file)
    lines: list[str] = []
    lines.append(f"# Survey: {Path(args.file).name}\n")
    lines.append(f"**Tabs**: {len(wb.sheetnames)}\n")

    total_cells = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = effective_row_count(ws)
        cols = effective_col_count(ws)
        total_cells += rows * cols

        lines.append(f"## Tab: {sheet_name}\n")
        lines.append(f"- **Rows**: {rows}")
        lines.append(f"- **Columns**: {cols}")

        # Merged cells
        merged = list(ws.merged_cells.ranges)
        if merged:
            lines.append(f"- **Merged cells**: {len(merged)} region(s)")

        # Frozen panes
        if ws.freeze_panes:
            lines.append(f"- **Frozen panes**: {ws.freeze_panes}")

        # Auto-filter
        if ws.auto_filter and ws.auto_filter.ref:
            lines.append(f"- **Auto-filter**: {ws.auto_filter.ref}")

        # Hidden sheet
        if ws.sheet_state != "visible":
            lines.append(f"- **Sheet state**: {ws.sheet_state}")

        # Preview: first 3 rows
        preview_rows = min(3, rows)
        if preview_rows > 0 and cols > 0:
            lines.append("")
            lines.append("**Preview (first 3 rows)**:\n")
            headers = [
                get_column_letter(c) for c in range(1, cols + 1)
            ]
            table_rows = []
            for r in range(1, preview_rows + 1):
                table_rows.append(
                    [cell_value(ws.cell(row=r, column=c)) for c in range(1, cols + 1)]
                )
            lines.append(md_table(headers, table_rows))

        lines.append("")

    lines.append(f"---\n**Total estimated cells**: {total_cells:,}")
    if total_cells > 50_000:
        lines.append(
            "\n**WARNING**: Cell count exceeds 50,000. Consider using `--tabs` "
            "to scope processing to specific tabs."
        )

    return truncate_output("\n".join(lines))


def cmd_headers(args: argparse.Namespace) -> str:
    """Extract column headers, sample rows, and inferred types."""
    wb = open_workbook(args.file)
    ws = get_sheet(wb, args.tab)
    rows = effective_row_count(ws)
    cols = effective_col_count(ws)

    if rows == 0 or cols == 0:
        return f"# Headers: {args.tab}\n\nTab is empty."

    lines: list[str] = []
    lines.append(f"# Headers: {args.tab}\n")

    # Detect header row: use row 1 by default, but if row 1 looks like a
    # title (single merged cell or single non-empty cell), try row 2.
    header_row = 1
    non_empty_r1 = sum(
        1 for c in range(1, cols + 1) if ws.cell(row=1, column=c).value is not None
    )
    if non_empty_r1 <= 1 and rows >= 2:
        non_empty_r2 = sum(
            1 for c in range(1, cols + 1) if ws.cell(row=2, column=c).value is not None
        )
        if non_empty_r2 > non_empty_r1:
            header_row = 2
            lines.append(f"*Header row detected at row {header_row} (row 1 appears to be a title)*\n")

    # Column headers
    col_headers = []
    for c in range(1, cols + 1):
        val = cell_value(ws.cell(row=header_row, column=c))
        letter = get_column_letter(c)
        col_headers.append(f"{letter}: {val}" if val else f"{letter}: *(empty)*")
    lines.append("**Columns**:\n")
    for h in col_headers:
        lines.append(f"- {h}")
    lines.append("")

    # Sample rows (up to 5 data rows after header)
    sample_start = header_row + 1
    sample_end = min(sample_start + 4, rows)
    if sample_start <= rows:
        lines.append(f"**Sample rows** ({sample_start}–{sample_end}):\n")
        headers = [cell_value(ws.cell(row=header_row, column=c)) or get_column_letter(c)
                    for c in range(1, cols + 1)]
        table_rows = []
        for r in range(sample_start, sample_end + 1):
            table_rows.append(
                [cell_value(ws.cell(row=r, column=c)) for c in range(1, cols + 1)]
            )
        lines.append(md_table(headers, table_rows))
        lines.append("")

    # Inferred column types (from first 20 data rows)
    lines.append("**Inferred column types** (from sample):\n")
    type_end = min(header_row + 20, rows)
    for c in range(1, cols + 1):
        values = []
        for r in range(header_row + 1, type_end + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                values.append(v)
        inferred = _infer_type(values)
        header = cell_value(ws.cell(row=header_row, column=c)) or get_column_letter(c)
        lines.append(f"- **{header}**: {inferred}")

    return truncate_output("\n".join(lines))


def _infer_type(values: list) -> str:
    """Infer a rough data type from a sample of values."""
    if not values:
        return "empty"
    types: Counter[str] = Counter()
    for v in values:
        if isinstance(v, bool):
            types["boolean"] += 1
        elif isinstance(v, int):
            types["integer"] += 1
        elif isinstance(v, float):
            types["decimal"] += 1
        elif hasattr(v, "strftime"):
            types["date/datetime"] += 1
        else:
            s = str(v).strip()
            if not s:
                types["empty"] += 1
            elif s.replace(".", "", 1).replace("-", "", 1).isdigit():
                types["numeric-string"] += 1
            else:
                types["text"] += 1
    non_empty = {k: v for k, v in types.items() if k != "empty"}
    if not non_empty:
        return "empty"
    dominant = max(non_empty, key=lambda k: non_empty[k])
    total = sum(non_empty.values())
    pct = non_empty[dominant] / total * 100
    nulls = types.get("empty", 0)
    suffix = f" ({nulls} null)" if nulls else ""
    if pct == 100:
        return f"{dominant}{suffix}"
    return f"{dominant} ({pct:.0f}%){suffix}"


def cmd_formatting(args: argparse.Namespace) -> str:
    """Extract formatting metadata: fills, conditional formatting, etc."""
    # Re-open without data_only to see formatting
    wb = openpyxl.load_workbook(args.file, read_only=False, data_only=False)
    ws = get_sheet(wb, args.tab)
    rows = effective_row_count(ws)
    cols = effective_col_count(ws)

    lines: list[str] = []
    lines.append(f"# Formatting: {args.tab}\n")

    # Conditional formatting rules
    cf_rules = list(ws.conditional_formatting)
    if cf_rules:
        lines.append(f"**Conditional formatting**: {len(cf_rules)} rule(s)\n")
        for i, cf in enumerate(cf_rules, 1):
            for rule in cf.rules:
                lines.append(f"- Rule {i}: type={rule.type}, range={cf.sqref}")
                if rule.formula:
                    lines.append(f"  Formula: {rule.formula}")
                if rule.dxf and rule.dxf.fill:
                    fg = rule.dxf.fill.fgColor
                    if fg and fg.rgb and fg.rgb != "00000000":
                        lines.append(f"  Fill: #{fg.rgb}")
        lines.append("")

    # Distinct fill colours and frequency
    fill_counter: Counter[str] = Counter()
    font_styles: Counter[str] = Counter()

    for r in range(1, min(rows + 1, 500)):  # cap at 500 rows for performance
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            # Fill
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                rgb = cell.fill.fgColor.rgb
                if isinstance(rgb, str) and rgb != "00000000":
                    fill_counter[rgb] += 1
            # Font styles
            if cell.font:
                if cell.font.bold:
                    font_styles["bold"] += 1
                if cell.font.strikethrough:
                    font_styles["strikethrough"] += 1
                if cell.font.italic:
                    font_styles["italic"] += 1

    if fill_counter:
        lines.append("**Distinct fill colours** (by cell count):\n")
        for colour, count in fill_counter.most_common(15):
            lines.append(f"- `#{colour}`: {count} cells")
        lines.append("")

    if font_styles:
        lines.append("**Font styles**:\n")
        for style, count in font_styles.most_common():
            lines.append(f"- {style}: {count} cells")
        lines.append("")

    # Hidden rows
    hidden_rows = [
        r for r in range(1, rows + 1)
        if ws.row_dimensions[r].hidden
    ]
    if hidden_rows:
        if len(hidden_rows) <= 20:
            lines.append(f"**Hidden rows**: {', '.join(str(r) for r in hidden_rows)}")
        else:
            lines.append(f"**Hidden rows**: {len(hidden_rows)} rows hidden")
        lines.append("")

    # Hidden columns
    hidden_cols = [
        get_column_letter(c) for c in range(1, cols + 1)
        if ws.column_dimensions[get_column_letter(c)].hidden
    ]
    if hidden_cols:
        lines.append(f"**Hidden columns**: {', '.join(hidden_cols)}")
        lines.append("")

    # Data validation rules
    validations = list(ws.data_validations.dataValidation) if ws.data_validations else []
    if validations:
        lines.append(f"**Data validation rules**: {len(validations)}\n")
        for dv in validations[:10]:
            lines.append(f"- Range: {dv.sqref}, Type: {dv.type}")
            if dv.formula1:
                lines.append(f"  Formula: {dv.formula1}")
        lines.append("")

    # Row groupings (outline levels)
    grouped_rows = [
        r for r in range(1, rows + 1)
        if ws.row_dimensions[r].outlineLevel and ws.row_dimensions[r].outlineLevel > 0
    ]
    if grouped_rows:
        lines.append(f"**Row groupings**: {len(grouped_rows)} rows with outline levels")
        lines.append("")

    if not fill_counter and not font_styles and not cf_rules and not hidden_rows and not hidden_cols:
        lines.append("No notable formatting detected.")

    return truncate_output("\n".join(lines))


def cmd_range(args: argparse.Namespace) -> str:
    """Extract cell values for a specified range."""
    wb = open_workbook(args.file)
    ws = get_sheet(wb, args.tab)
    rows = effective_row_count(ws)
    cols = effective_col_count(ws)

    # Parse row range
    row_start, row_end = 1, rows
    if args.rows:
        parts = args.rows.split(":")
        if len(parts) == 2:
            row_start = int(parts[0])
            row_end = int(parts[1])
        else:
            print(f"Error: invalid row range '{args.rows}'. Use START:END.", file=sys.stderr)
            sys.exit(1)

    # Parse column range
    col_start, col_end = 1, cols
    if args.cols:
        parts = args.cols.split(":")
        if len(parts) == 2:
            col_start = column_index_from_string(parts[0].upper())
            col_end = column_index_from_string(parts[1].upper())
        else:
            print(f"Error: invalid column range '{args.cols}'. Use A:H.", file=sys.stderr)
            sys.exit(1)

    # Clamp
    row_start = max(1, row_start)
    row_end = min(row_end, rows)
    col_start = max(1, col_start)
    col_end = min(col_end, cols)

    lines: list[str] = []
    lines.append(f"# Range: {args.tab} [rows {row_start}:{row_end}, cols {get_column_letter(col_start)}:{get_column_letter(col_end)}]\n")

    # Use row 1 (or detected header row) as column headers
    headers = [
        cell_value(ws.cell(row=1, column=c)) or get_column_letter(c)
        for c in range(col_start, col_end + 1)
    ]

    table_rows = []
    for r in range(row_start, row_end + 1):
        table_rows.append(
            [cell_value(ws.cell(row=r, column=c)) for c in range(col_start, col_end + 1)]
        )

    lines.append(md_table(headers, table_rows))
    lines.append(f"\n**Rows returned**: {len(table_rows)}")

    return truncate_output("\n".join(lines))


def cmd_lookup(args: argparse.Namespace) -> str:
    """Extract full content of a small reference/lookup tab."""
    wb = open_workbook(args.file)
    ws = get_sheet(wb, args.tab)
    rows = effective_row_count(ws)
    cols = effective_col_count(ws)
    max_rows = args.max_rows

    if rows == 0 or cols == 0:
        return f"# Lookup: {args.tab}\n\nTab is empty."

    lines: list[str] = []
    lines.append(f"# Lookup: {args.tab}\n")

    capped = min(rows, max_rows + 1)  # +1 for header row
    if rows > max_rows + 1:
        lines.append(
            f"**WARNING**: Tab has {rows - 1} data rows; showing first {max_rows} "
            f"(capped by --max-rows).\n"
        )

    # Header row
    headers = [
        cell_value(ws.cell(row=1, column=c)) or get_column_letter(c)
        for c in range(1, cols + 1)
    ]

    table_rows = []
    for r in range(2, capped + 1):
        table_rows.append(
            [cell_value(ws.cell(row=r, column=c)) for c in range(1, cols + 1)]
        )

    lines.append(md_table(headers, table_rows))
    lines.append(f"\n**Rows**: {len(table_rows)} (of {rows - 1} total)")

    return truncate_output("\n".join(lines))


# ── CLI ──────────────────────────────────────────────────────────────

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="excel_tool",
        description="Excel interrogation tool for the Excel-to-Satsuma skill.",
    )
    sub = parser.add_subparsers(dest="command", required=True)

    # survey
    p_survey = sub.add_parser("survey", help="Survey workbook structure")
    p_survey.add_argument("file", help="Path to .xlsx file")

    # headers
    p_headers = sub.add_parser("headers", help="Extract column headers and sample rows")
    p_headers.add_argument("file", help="Path to .xlsx file")
    p_headers.add_argument("tab", help="Tab/sheet name")

    # formatting
    p_fmt = sub.add_parser("formatting", help="Extract formatting metadata")
    p_fmt.add_argument("file", help="Path to .xlsx file")
    p_fmt.add_argument("tab", help="Tab/sheet name")

    # range
    p_range = sub.add_parser("range", help="Extract cell values for a range")
    p_range.add_argument("file", help="Path to .xlsx file")
    p_range.add_argument("tab", help="Tab/sheet name")
    p_range.add_argument("--rows", help="Row range, e.g. 2:50")
    p_range.add_argument("--cols", help="Column range, e.g. A:H")

    # lookup
    p_lookup = sub.add_parser("lookup", help="Extract full lookup tab content")
    p_lookup.add_argument("file", help="Path to .xlsx file")
    p_lookup.add_argument("tab", help="Tab/sheet name")
    p_lookup.add_argument("--max-rows", type=int, default=500, help="Max data rows (default 500)")

    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    dispatch = {
        "survey": cmd_survey,
        "headers": cmd_headers,
        "formatting": cmd_formatting,
        "range": cmd_range,
        "lookup": cmd_lookup,
    }

    output = dispatch[args.command](args)
    print(output)


if __name__ == "__main__":
    main()
